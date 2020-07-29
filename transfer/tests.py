from django.test import TestCase
from openpyxl import load_workbook

# Create your tests here

wb_obj = load_workbook(filename='/home/unhmguest/Desktop/transfereval-docs/transfers_bymajor.xlsx', data_only=True)
print(wb_obj.sheetnames)


class CheckEvaluationTest(TestCase):
    def setup(self):
        CheckEvaluation.objects.create(
                                      school_name='University at Manchester',
                                      major_name='Computing technology',
                                      transfer_course_name='CMP 555',
                                      transfer_course_title='Computing for data science',
                                      unhm_equivalent='COMP 881',
                                      approver_name='Sabin',
                                      approved_status='Yes',
                                      sem_or_year_taken='SP 17',
                                      expiration_date='22-05-2017'
                                      )

    def (self):
